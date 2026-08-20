import streamlit as st
import pandas as pd
import io
import openpyxl
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from leads_scraper import run_lead_hunter

st.set_page_config(page_title="Lead Hunter Portal", page_icon="🌴", layout="wide")

# Force Pure Light / White Theme Only & Remove Header Elements
st.markdown("""
    <style>
    /* تثبيت اللون الأبيض لجميع خلفيات التطبيق بدون استثناء */
    html, body, [data-testid="stAppViewContainer"], .stApp {
        background-color: #FFFFFF !important;
        color: #000000 !important;
    }
    
    [data-testid="stHeader"] {
        background-color: #FFFFFF !important;
    }

    .main-title {
        color: #1b365d;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 700;
        text-align: center;
        margin-bottom: 0px;
    }
    .sub-title {
        color: #666666;
        font-size: 13px;
        text-align: center;
        letter-spacing: 1px;
        margin-bottom: 25px;
    }
    .stTextInput>div>div>input {
        border: 1px solid #1b365d !important;
        border-radius: 6px;
        background-color: #FFFFFF !important;
        color: #000000 !important;
    }
    [data-testid="stForm"] {
        border: none !important;
        padding: 0px !important;
        background-color: #FFFFFF !important;
    }
    .stButton>button, .stFormSubmitButton>button {
        background-color: #1b365d !important;
        color: white !important;
        border-radius: 6px;
        font-weight: 600;
        height: 45px;
        width: 100%;
    }
    div[data-testid="stColumn"] > div {
        background-color: #FFFFFF !important;
    }
    </style>
""", unsafe_allow_html=True)

# --- نظام التراخيص والتحكم في الاشتراك (Subscription Control) ---
CLIENT_LICENSE = {
    "company_name": "Lead Portal",
    "status": "Active",            # غيرها لـ "Disabled" عشان تقفل عليهم السيستم فوراً
    "expiry_date": "2026-12-31",   # تاريخ انتهاء الاشتراك (YYYY-MM-DD)
    "admin_user": "admin",
    "admin_pass": "2026"           # تم التحديث إلى 2026
}

def check_subscription():
    """فحص حالة الاشتراك والتاريخ"""
    if CLIENT_LICENSE["status"] != "Active":
        return False, "Your account subscription is currently suspended. Please contact support."
    
    expiry_dt = datetime.strptime(CLIENT_LICENSE["expiry_date"], "%Y-%m-%d")
    if datetime.now() > expiry_dt:
        return False, f"Your subscription expired on {CLIENT_LICENSE['expiry_date']}. Please renew to access the portal."
    
    return True, "OK"

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

def login():
    col1, col2, col3 = st.columns([1.2, 2, 1.2])
    with col2:
        st.markdown("<h2 class='main-title'>Lead Hunter Portal</h2>", unsafe_allow_html=True)
        st.markdown("<p class='sub-title'>AUTOMATED B2B LEAD EXTRACTION & VERIFICATION ENGINE</p>", unsafe_allow_html=True)
        
        with st.form("login_form", clear_on_submit=False):
            username = st.text_input("Username", placeholder="Enter username")
            password = st.text_input("Password", type="password", placeholder="••••••••")
            submit_button = st.form_submit_button("Sign In to Portal", use_container_width=True)
            
            if submit_button:
                # 1. التحقق من حالة الاشتراك الأول
                is_valid_license, license_msg = check_subscription()
                
                if not is_valid_license:
                    st.error(f"⛔ Access Denied: {license_msg}")
                elif username == CLIENT_LICENSE["admin_user"] and password == CLIENT_LICENSE["admin_pass"]:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("Invalid username or password")

def generate_colored_excel(df):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validated Leads"

    headers = list(df.columns)
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, row_data in enumerate(df.values, start=2):
        ws.append(list(row_data))
        status_val = str(row_data[1])
        
        status_cell = ws.cell(row=row_idx, column=2)
        if status_val == "High Confidence":
            status_cell.fill = green_fill
        elif status_val == "Medium Confidence":
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = orange_fill
            
        status_cell.alignment = Alignment(horizontal="center")

    wb.save(output)
    return output.getvalue()

if not st.session_state.authenticated:
    login()
else:
    col_head, col_out = st.columns([5, 1])
    with col_head:
        st.markdown("<h2 style='color: #1b365d; margin: 0;'>Lead Hunter Portal</h2>", unsafe_allow_html=True)
        st.caption("Automated Corporate Lead Verification & Cost Optimization Engine")
    with col_out:
        if st.button("Sign Out"):
            st.session_state.authenticated = False
            st.rerun()

    st.divider()

    uploaded_file = st.file_uploader("Upload Companies Sheet (CSV or Excel)", type=["csv", "xlsx"])

    if uploaded_file:
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
            
        st.subheader("📊 Raw Sheet Preview")
        st.dataframe(df.head(), use_container_width=True)
        
        column_name = st.selectbox("Select Company Name Column:", df.columns)
        total_rows = len(df[column_name].dropna().tolist())
        
        st.divider()
        st.subheader("⚙️ Execution Range Settings (Cost Management)")
        
        selected_range = st.slider(
            "Select Company Range via Slider:",
            min_value=1,
            max_value=total_rows,
            value=(1, min(20, total_rows))
        )
        
        c1, c2 = st.columns(2)
        with c1:
            start_row = st.number_input("Or Enter Start Row Number:", min_value=1, max_value=total_rows, value=selected_range[0])
        with c2:
            end_row = st.number_input("Or Enter End Row Number:", min_value=start_row, max_value=total_rows, value=selected_range[1])
            
        selected_count = int(end_row - start_row + 1)
        st.info(f"Targeting **{selected_count}** companies (From row {int(start_row)} to {int(end_row)}). Duplicate companies will load from local Cache instantly without API consumption.")
        
        if st.button(f"🚀 Start Extraction ({selected_count} Leads)", use_container_width=True):
            companies_list = df[column_name].dropna().tolist()
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            def update_status(current, total, name):
                progress_bar.progress(current / total)
                status_text.text(f"Processing ({current}/{total}): {name}")
                
            results = run_lead_hunter(companies_list, start_idx=int(start_row), end_idx=int(end_row), status_callback=update_status)
            res_df = pd.DataFrame(results)
            
            st.success("✅ Extraction & Verification Completed Successfully!")
            st.subheader("📋 Final Verified Leads Table")
            st.dataframe(res_df, use_container_width=True)
            
            excel_colored_data = generate_colored_excel(res_df)
            
            st.download_button(
                label="📥 Download Color-Coded Excel Sheet (.xlsx)",
                data=excel_colored_data,
                file_name="Verified_Corporate_Leads.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )